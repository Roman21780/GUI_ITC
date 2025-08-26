from openpyxl import load_workbook
from db_access import AccessDatabase
import logging
import datetime

# Настройка логирования
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def migrate_excel_to_access():
    """Переносит данные из Excel в Access"""
    try:
        # Загружаем Excel-файл
        file_path = 'C:\\Work\\GUI_ITC\\Report.xlsx'
        wb = load_workbook(file_path)
        ws = wb['current']

        # Создаем экземпляр базы данных
        db = AccessDatabase()

        # Список для хранения записей
        records = []

        # Читаем данные из Excel
        for row in ws.iter_rows(min_row=2, values_only=True):  # Пропускаем заголовок (первая строка)
            record = {
                'company': row[0],
                'field': row[1],
                'well': row[2],
                'date_research': row[3],  # Убедитесь, что это дата
                'created_date': datetime.datetime.now()  # Текущая дата создания записи
            }
            records.append(record)

        # Вставляем данные в таблицу main_data
        for record in records:
            try:
                db.insert_main_data(record)
                logging.info(f"Запись успешно добавлена: {record}")
            except Exception as e:
                logging.error(f"Ошибка при добавлении записи {record}: {str(e)}")

        # Закрываем соединение с базой данных
        db.close()

        logging.info("Миграция данных завершена успешно")

    except FileNotFoundError:
        logging.error(f"Файл Excel не найден: {file_path}")
    except KeyError:
        logging.error("Лист 'current' не найден в файле Excel")
    except Exception as e:
        logging.error(f"Ошибка миграции: {str(e)}")

if __name__ == "__main__":
    migrate_excel_to_access()